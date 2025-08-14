# COOL_CRAFT_WEBAPP_FULL.py
import streamlit as st
import pandas as pd
import os
import io
import re
from math import ceil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell

# ================ PAGE SETUP ================
st.set_page_config(page_title="CoolCraft VRF TDS Generator", layout="wide")

# ================ CSS ================
st.markdown("""
<style>
.header {font-size:28px; font-weight:700; color:#fff; text-align:center; padding:14px;
 background:linear-gradient(90deg,#0047ab,#e63946);}
.card {background:#fff; padding:14px; border-radius:10px; box-shadow:0 1px 12px rgba(0,0,0,0.15); margin-bottom:12px;}
.monitor {background:#f1f8ff; padding:10px; border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,0.06);}
</style>
""", unsafe_allow_html=True)
st.markdown('<div class="header">❄️ CoolCraft TDS – HVAC Technical Data Sheet Generator</div>', unsafe_allow_html=True)

# ================ DATA SOURCES ================
DATA_SOURCES = {
    ("Toshiba", "VRF", "Other", "Outdoor Unit", "Single Unit"): "data/TOS_VRF_SINGLE.xlsx",
    ("Toshiba", "VRF", "Other", "Outdoor Unit", "High Efficiency"): "data/TOS_VRF_HI_EFM.xlsx",
    ("Toshiba", "VRF", "Other", "Outdoor Unit", "Combination"): "data/TOS_VRF_COMB.xlsx",
    ("Toshiba", "VRF", "Cassette", "Indoor Unit", None): "data/4-way Cassette U series.xlsx",
    ("Toshiba", "VRF", "High Wall", "Indoor Unit", None): "data/Highwall_USeries.xlsx",
    ("Toshiba", "VRF", "Ductable", "Indoor Unit", None): "data/HS Ductable TDS.xlsx",
}

# Conversion constants
KW_TO_HP = 1.0 / 0.745699872
TON_TO_HP = 3.517 / 0.745699872

# ================ HELPERS ================
@st.cache_data(ttl=600)
def load_excel_all_sheets(path: str):
    if os.path.exists(path):
        xls = pd.ExcelFile(path)
    else:
        raise FileNotFoundError(f"File not found: {path}")
    return {sheet: pd.read_excel(xls, sheet_name=sheet) for sheet in xls.sheet_names}

def normalize_name(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('\u200b', '')
    return s

def build_normalized_map(df: pd.DataFrame) -> dict:
    return {orig: normalize_name(orig) for orig in df.columns}

def find_capacity_column_by_type(df: pd.DataFrame, unit_type: str) -> str:
    norm_map = build_normalized_map(df)
    if unit_type.lower() == "indoor":
        for orig, norm in norm_map.items():
            if "cooling capacity" in norm and "kw" in norm:
                return orig
        for orig, norm in norm_map.items():
            if "capacity" in norm and "kw" in norm:
                return orig
        for orig, norm in norm_map.items():
            if "kw" in norm:
                return orig
    else:
        for orig, norm in norm_map.items():
            if "hp" in norm and ("capacity" in norm or "hp" in norm):
                return orig
        for orig, norm in norm_map.items():
            if "horsepower" in norm or re.search(r'\bhp\b', norm):
                return orig
        for orig, norm in norm_map.items():
            if "capacity" in norm and ("hp" in norm or "horsepower" in norm):
                return orig
    return None

def expand_combo_instances(combo):
    inst = []
    for hp, cnt in sorted(combo.items(), reverse=True):
        inst.extend([hp] * cnt)
    return inst

def greedy_combo(target_cap, sizes):
    rem = int(round(target_cap))
    combo = {}
    for s in sizes:
        cnt = rem // s
        if cnt > 0:
            combo[s] = int(cnt)
            rem -= s * cnt
    if rem > 0 and sizes:
        smallest = min(sizes)
        combo[smallest] = combo.get(smallest, 0) + 1
    return combo

def generate_candidate_combos(target_cap, sizes, max_suggestions=12):
    raw = [greedy_combo(target_cap, sizes)]
    for s in sizes[:6]:
        raw.append({s: ceil(target_cap / s)})
    uniq = {tuple(sorted(c.items(), reverse=True)): c for c in raw}
    combos = list(uniq.values())[:max_suggestions]
    scored = []
    for c in combos:
        total = sum(expand_combo_instances(c))
        units = sum(c.values())
        closeness = 1.0 / (1 + abs(total - target_cap) / max(1, target_cap))
        unit_score = 1.0 / (1 + units)
        score = 0.6 * closeness + 0.4 * unit_score
        scored.append((score, c))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [c for _, c in scored]

def find_nearest_row(df, target_val, cap_col):
    diffs = (pd.to_numeric(df[cap_col], errors='coerce') - float(target_val)).abs()
    idx = diffs.idxmin()
    return df.loc[idx] if pd.notna(idx) else None

def export_excel(df: pd.DataFrame, sheet_name='VRF_TDS_Report'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True, color='FFFFFF')
    col_count = df.shape[1] if df.shape[1] > 0 else 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="❄️ CoolCraft Technical Data Sheet (TDS)")
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = center_align
    title_cell.fill = PatternFill(start_color='4B8BBE', end_color='4B8BBE', fill_type='solid')
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 2:
                cell.font = header_font
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                cell.alignment = center_align
            else:
                fill_color = "EAF1FB" if r_idx % 2 == 0 else "FFFFFF"
                if isinstance(value, (int, float)):
                    if value >= 90:
                        fill_color = "1e824c"
                    elif value >= 70:
                        fill_color = "f39c12"
                    else:
                        fill_color = "c0392b"
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.border = border_all
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        first_cell = next((cell for cell in col_cells if not isinstance(cell, MergedCell)), None)
        if first_cell:
            try:
                ws.column_dimensions[first_cell.column_letter].width = min(length + 3, 50)
            except Exception:
                pass
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ================ SIDEBAR / WIZARD ================
st.sidebar.header("Start New TDS")
brand = st.sidebar.selectbox("Brand", ["Toshiba", "Carrier"])
system_type = st.sidebar.selectbox("System Type", ["VRF", "Non-VRF"])
product_type = st.sidebar.selectbox("Product Type", ["Cassette", "High Wall", "Ductable", "Other"])
unit_type = st.sidebar.selectbox("Unit Type", ["Indoor", "Outdoor"])
combination_mode = st.sidebar.selectbox("Combination Mode", ["Automatic", "Manual"])
combination_type = None
if unit_type == "Outdoor" and product_type == "Other":
    combination_type = st.sidebar.selectbox("Outdoor selection", ["Combination", "High Efficiency", "Single Unit"])
if st.sidebar.button("Proceed"):
    st.session_state['wizard'] = {
        'brand': brand, 'system_type': system_type, 'product_type': product_type,
        'unit_type': unit_type, 'combination_mode': combination_mode,
        'combination_type': combination_type
    }
wizard = st.session_state.get('wizard')
if not wizard:
    st.stop()

# ================ LOAD DATA & PREVIEW ================
def map_key(w):
    if w['unit_type'] == "Outdoor" and w['product_type'] == "Other":
        return (w['brand'], w['system_type'], "Other", "Outdoor Unit", w.get('combination_type'))
    else:
        return (w['brand'], w['system_type'], w['product_type'], "Indoor Unit", None)

path = DATA_SOURCES.get(map_key(wizard))
if not path:
    st.error("No mapping found for this selection.")
    st.stop()

sheets = load_excel_all_sheets(path)
sheet_choice = st.selectbox("Select sheet", list(sheets)) if len(sheets) > 1 else list(sheets)[0]
df = sheets[sheet_choice].copy()
cap_col = find_capacity_column_by_type(df, wizard['unit_type'])
if wizard['unit_type'] == "Outdoor": cap_label_type = "HP"
else: cap_label_type = "kW"
if cap_col is None:
    sizes_available = []
else:
    df[cap_col] = pd.to_numeric(df[cap_col], errors='coerce')
    sizes_available = sorted(list({float(x) for x in df[cap_col].dropna().unique()}))

st.subheader("📊 Dataset Preview")
st.dataframe(df.head(10))

# ================ TWO-COLUMN DASHBOARD ================
col1, col2 = st.columns([1,1])

with col1:
    st.markdown("### ⚡ Monitors / Info Panel")
    st.markdown('<div class="monitor">Loaded rows: <b>{}</b></div>'.format(len(df)), unsafe_allow_html=True)
    st.markdown('<div class="monitor">Available capacities: <b>{}</b></div>'.format(sizes_available), unsafe_allow_html=True)
    st.markdown('<div class="monitor">Capacity Column: <b>{}</b></div>'.format(cap_col), unsafe_allow_html=True)

with col2:
    st.subheader("💡 TDS Preview & Ratings")
    # Placeholder; actual preview filled after combination generation

# ================ COMBINATION GENERATION & TDS ================
if wizard['combination_mode'] == "Automatic" and sizes_available:
    target_cap = st.number_input(f"Enter Target Capacity ({cap_label_type})", min_value=1.0, value=float(sizes_available[-1]*2))
    if st.button("Generate Combinations"):
        combos = generate_candidate_combos(target_cap, sizes_available)
        enriched = []
        for combo in combos:
            rows = []
            for idx, cap in enumerate(expand_combo_instances(combo)):
                exact = df[pd.to_numeric(df[cap_col], errors='coerce') == float(cap)]
                chosen = exact.iloc[0].to_dict() if not exact.empty else find_nearest_row(df, cap, cap_col).to_dict()
                chosen['_instance'] = idx+1
                rows.append(chosen)
            enriched.append({'combo': combo, 'rows': rows, 'total_cap': sum(expand_combo_instances(combo)), 'units': len(rows)})
        st.session_state['enriched'] = enriched

elif wizard['combination_mode'] == "Manual":
    man_in = st.text_input(f"Enter {cap_label_type} sizes (use +, e.g. 3.5+3.5+2)", "")
    if st.button("Create Combo") and man_in.strip():
        sizes = [float(x) for x in man_in.split("+") if x.strip()]
        rows = []
        for idx, cap in enumerate(sizes):
            nearest = find_nearest_row(df, cap, cap_col) if cap_col else {}
            chosen = nearest.to_dict() if nearest is not None else {'_cap_input': cap}
            chosen['_instance'] = idx+1
            rows.append(chosen)
        combo_dict = {}
        for s in sizes: key = int(s) if float(s).is_integer() else s; combo_dict[key]=combo_dict.get(key,0)+1
        st.session_state['enriched'] = [{'combo': combo_dict, 'rows': rows, 'total_cap': sum(sizes), 'units': len(rows)}]

# ================ SHOW & EXPORT TDS ================
if 'enriched' in st.session_state:
    enriched = st.session_state['enriched']
    for i, e in enumerate(enriched, 1):
        desc = " + ".join(f"{v}×{k}{cap_label_type}" for k, v in e['combo'].items())
        st.markdown(f"**Option {i}**: {desc} — Units: {e.get('units',0)} — Total: {round(e.get('total_cap',0),3)} {cap_label_type}")

    choice = st.selectbox("Choose option", range(1, len(enriched)+1)) - 1
    chosen = enriched[choice]

    mod_rows = []
    for r in chosen['rows']:
        inst = int(r.get('_instance',0))
        sel_row = r.copy()
        cap_value = r.get(cap_col,r.get('_cap_input','N/A'))
        st.markdown(f"Instance {inst} — {cap_label_type}: {cap_value}")
        mod_rows.append(sel_row)

    out_df = pd.DataFrame(mod_rows)
    # metadata
    client = st.text_input("Client Name", "Client")
    manuf = st.text_input("Manufacturer", wizard['brand'])
    billing = st.text_input("Billing/Sales", "")
    rdate = st.date_input("Report Date", datetime.now().date())
    out_df['Client'] = client
    out_df['Manufacturer'] = manuf
    out_df['BillingSales'] = billing
    out_df['ReportDate'] = rdate.strftime('%Y-%m-%d')
    out_df['SelectedCombo'] = " + ".join(f"{v}×{k}{cap_label_type}" for k,v in chosen['combo'].items())
    out_df[f'ComboTotal{cap_label_type}'] = chosen.get('total_cap',0)
    out_df['ComboUnits'] = chosen.get('units',0)

    st.subheader("✅ Final TDS Preview")
    num_cols = list(out_df.select_dtypes(include=['number']).columns)
    def style_func(x):
        if isinstance(x,(int,float)):
            if x>=90: return 'background-color: #1e824c; color:white'
            elif x>=70: return 'background-color: #f39c12; color:white'
            else: return 'background-color: #c0392b; color:white'
        return ''
    if num_cols: st.dataframe(out_df.style.applymap(style_func, subset=num_cols))
    else: st.dataframe(out_df)

    if st.button("Download Excel"):
        bytes_xl = export_excel(out_df)
        fname = f"TDS_{client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button("📥 Download TDS Excel", data=bytes_xl, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
